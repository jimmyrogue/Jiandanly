"""Excel-specific office.* outline, read, and copy-on-write tools."""

from __future__ import annotations

import io
import re
from typing import Any

from langchain_core.tools import tool
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import column_index_from_string
from openpyxl.utils import get_column_letter as get_column_letter
from openpyxl.utils.cell import range_boundaries

from .office_common import atomic_write as _atomic_write
from .office_common import ensure_copy_for_write as _ensure_copy_for_write
from .office_common import resolve_read_source as _resolve_read_source
from .office_common import validate_path as _validate_path
from .office_common import write_error as _write_error
from .office_common import write_result as _write_result


def _outline_xlsx(path: Any) -> dict[str, Any]:
    """Cheap structural summary of a .xlsx — sheet names + dimensions."""
    wb = load_workbook(path, read_only=True, data_only=True)
    sheets: list[dict[str, Any]] = []
    try:
        for name in wb.sheetnames:
            ws = wb[name]
            sheets.append(
                {
                    "name": name,
                    "rows": ws.max_row or 0,
                    "columns": ws.max_column or 0,
                }
            )
    finally:
        wb.close()
    return {"sheets": sheets}


# ═══════════════════════════════════════════════════════════════════════
# XLSX write helpers
# ═══════════════════════════════════════════════════════════════════════


_CELL_REF_RE = re.compile(r"^([A-Za-z]+)(\d+)$")


def _parse_cell_ref(ref: str) -> tuple[int, int]:
    """Parse an A1-style cell ref into (row, col) 1-indexed tuple."""
    m = _CELL_REF_RE.match(ref.strip())
    if not m:
        raise ValueError(f"invalid cell reference: {ref!r} (expected like 'A1')")
    col_str, row_str = m.groups()
    return int(row_str), column_index_from_string(col_str.upper())


def _parse_range(range_ref: str) -> tuple[int, int, int, int]:
    """Parse a range like 'A1:C3' (or 'A1' for a single cell) into
    (min_row, min_col, max_row, max_col), all 1-indexed."""
    ref = range_ref.strip()
    if ":" not in ref:
        # Single cell — treat as a 1x1 range.
        row, col = _parse_cell_ref(ref)
        return row, col, row, col
    min_col, min_row, max_col, max_row = range_boundaries(ref.upper())
    return min_row, min_col, max_row, max_col


def _resolve_sheet(wb, sheet: str | None):
    """Return the worksheet by name, or the active sheet when sheet is None."""
    if not sheet:
        return wb.active
    if sheet not in wb.sheetnames:
        raise ValueError(f"sheet not found: {sheet!r}; available: {wb.sheetnames}")
    return wb[sheet]


def _color_to_hex(color: str) -> str:
    """Normalize a CSS-ish color string into openpyxl's 8-char AARRGGBB
    or 6-char RRGGBB hex (without leading '#'). Accepts '#RRGGBB',
    'RRGGBB', '#AARRGGBB', 'AARRGGBB'. Bare names not supported."""
    c = color.strip().lstrip("#").upper()
    if len(c) not in (6, 8) or not all(ch in "0123456789ABCDEF" for ch in c):
        raise ValueError(f"color must be hex like '#FF5722' or '#80FF5722', got {color!r}")
    return c


# ═══════════════════════════════════════════════════════════════════════
# XLSX write tools
# ═══════════════════════════════════════════════════════════════════════


@tool("office.set_cells")
def office_set_cells(
    path: str,
    sheet: str | None,
    range: str,
    values: list[list[Any]],
) -> dict[str, Any]:
    """Write a 2D values block into a rectangular cell range.

    Writes to `<basename>.edited.xlsx`; original preserved.

    Args:
        path: .xlsx to edit.
        sheet: sheet name (or None for the active sheet).
        range: A1-style range like "A1:C3" or single cell "B5".
        values: 2D list shaped to match the range. Top-left is range
                top-left. Each inner list is one row. Extra cells in
                the range that lack values stay unchanged; extra
                values beyond the range are ignored.

    Returns:
        dict with `ok`, `original_path`, `edited_path`, `kind`,
        `summary={range, rows_written, cols_written, cells_written}`.
    """
    resolved, kind, err = _validate_path(path)
    if err is not None:
        return _write_error(None, err)
    assert resolved is not None and kind is not None
    if kind != "excel":
        return _write_error(kind, "office.set_cells requires a .xlsx file", resolved)
    if not values or not isinstance(values, list):
        return _write_error(kind, "values must be a non-empty 2D list", resolved)

    target_path = _ensure_copy_for_write(resolved)

    def _do_write(tmp_path: str) -> dict[str, Any]:
        wb = load_workbook(target_path)
        try:
            ws = _resolve_sheet(wb, sheet)
            min_row, min_col, max_row, max_col = _parse_range(range)
            rng_rows = max_row - min_row + 1
            rng_cols = max_col - min_col + 1
            written = 0
            for r_offset, row_values in enumerate(values):
                if r_offset >= rng_rows:
                    break
                if not isinstance(row_values, list):
                    raise ValueError(f"values[{r_offset}] must be a list")
                for c_offset, cell_value in enumerate(row_values):
                    if c_offset >= rng_cols:
                        break
                    ws.cell(row=min_row + r_offset, column=min_col + c_offset, value=cell_value)
                    written += 1
            wb.save(tmp_path)
        finally:
            wb.close()
        return {
            "range": range,
            "rows_written": min(len(values), rng_rows),
            "cols_written": rng_cols,
            "cells_written": written,
        }

    try:
        summary = _atomic_write(target_path, kind, _do_write)
    except Exception as exc:
        return _write_error(kind, f"write failed: {exc.__class__.__name__}: {exc}", resolved)
    return _write_result(resolved, target_path, kind, summary)


@tool("office.set_formula")
def office_set_formula(
    path: str,
    sheet: str | None,
    cell: str,
    formula: str,
) -> dict[str, Any]:
    """Write a formula into one cell. Writes to `<basename>.edited.xlsx`.

    NOTE: openpyxl writes the formula text — it does NOT evaluate it.
    The cell's displayed value updates only when Microsoft Excel /
    LibreOffice / Numbers opens the file. The right-side preview shows
    the literal formula text until then. This is an openpyxl limitation,
    not a bug in this tool.

    Args:
        path: .xlsx to edit.
        sheet: sheet name (None = active).
        cell: target cell like "D2".
        formula: must start with "=", e.g. "=SUM(A2:A10)",
                 "=IF(B2>0, \\"Pos\\", \\"Neg\\")", "=VLOOKUP(...)".

    Returns:
        dict with `ok`, `original_path`, `edited_path`, `kind`,
        `summary={cell, formula}`.
    """
    resolved, kind, err = _validate_path(path)
    if err is not None:
        return _write_error(None, err)
    assert resolved is not None and kind is not None
    if kind != "excel":
        return _write_error(kind, "office.set_formula requires a .xlsx file", resolved)
    formula = (formula or "").strip()
    if not formula.startswith("="):
        return _write_error(kind, "formula must start with '='", resolved)

    target_path = _ensure_copy_for_write(resolved)

    def _do_write(tmp_path: str) -> dict[str, Any]:
        wb = load_workbook(target_path)
        try:
            ws = _resolve_sheet(wb, sheet)
            row, col = _parse_cell_ref(cell)
            ws.cell(row=row, column=col, value=formula)
            wb.save(tmp_path)
        finally:
            wb.close()
        return {"cell": cell, "formula": formula}

    try:
        summary = _atomic_write(target_path, kind, _do_write)
    except Exception as exc:
        return _write_error(kind, f"write failed: {exc.__class__.__name__}: {exc}", resolved)
    return _write_result(resolved, target_path, kind, summary)


@tool("office.set_cell_format")
def office_set_cell_format(
    path: str,
    sheet: str | None,
    range: str,
    bold: bool | None = None,
    italic: bool | None = None,
    font_size: float | None = None,
    font_color: str | None = None,
    bg_color: str | None = None,
    align: str | None = None,
    border: bool | None = None,
) -> dict[str, Any]:
    """Apply font / fill / alignment / border formatting to a cell range.

    Writes to `<basename>.edited.xlsx`; original preserved. Only the
    arguments you pass are applied — None means "leave unchanged".

    Args:
        path: .xlsx to edit.
        sheet: sheet name (None = active).
        range: A1-style range.
        bold, italic: True / False (None to leave alone).
        font_size: e.g. 14.
        font_color, bg_color: hex like "#FF5722" or "FF5722". Alpha
                              prefix is accepted ("#80FF5722").
        align: one of "left" / "center" / "right" (horizontal only).
        border: True adds a thin black border to every cell in the
                range; False clears borders.

    Returns:
        dict with `ok`, `original_path`, `edited_path`, `kind`,
        `summary={range, cells_formatted, applied}` where `applied`
        is the list of attribute names that were actually changed.
    """
    resolved, kind, err = _validate_path(path)
    if err is not None:
        return _write_error(None, err)
    assert resolved is not None and kind is not None
    if kind != "excel":
        return _write_error(kind, "office.set_cell_format requires a .xlsx file", resolved)

    applied: list[str] = []
    try:
        font_color_hex = _color_to_hex(font_color) if font_color else None
        bg_color_hex = _color_to_hex(bg_color) if bg_color else None
    except ValueError as exc:
        return _write_error(kind, str(exc), resolved)
    if align is not None and align not in {"left", "center", "right"}:
        return _write_error(kind, f"align must be left/center/right, got {align!r}", resolved)

    target_path = _ensure_copy_for_write(resolved)

    def _do_write(tmp_path: str) -> dict[str, Any]:
        wb = load_workbook(target_path)
        try:
            ws = _resolve_sheet(wb, sheet)
            min_row, min_col, max_row, max_col = _parse_range(range)
            count = 0
            side = Side(border_style="thin", color="FF000000") if border else None
            border_style = (
                Border(left=side, right=side, top=side, bottom=side)
                if border
                else (Border() if border is False else None)
            )
            for r in range_iter(min_row, max_row):
                for c in range_iter(min_col, max_col):
                    cell_obj = ws.cell(row=r, column=c)
                    # Font — clone current font then patch fields the
                    # caller asked to change.
                    if any(v is not None for v in (bold, italic, font_size, font_color_hex)):
                        existing = cell_obj.font
                        cell_obj.font = Font(
                            name=existing.name,
                            size=font_size if font_size is not None else existing.size,
                            bold=bold if bold is not None else existing.bold,
                            italic=italic if italic is not None else existing.italic,
                            color=font_color_hex if font_color_hex is not None else existing.color,
                        )
                        for f in ("bold", "italic", "font_size", "font_color"):
                            if locals().get(f) is not None or (
                                f == "font_color" and font_color_hex
                            ):
                                if f not in applied:
                                    applied.append(f)
                    if bg_color_hex is not None:
                        cell_obj.fill = PatternFill(
                            fill_type="solid", start_color=bg_color_hex, end_color=bg_color_hex
                        )
                        if "bg_color" not in applied:
                            applied.append("bg_color")
                    if align is not None:
                        existing = cell_obj.alignment
                        cell_obj.alignment = Alignment(
                            horizontal=align,
                            vertical=existing.vertical,
                            wrap_text=existing.wrap_text,
                        )
                        if "align" not in applied:
                            applied.append("align")
                    if border_style is not None:
                        cell_obj.border = border_style
                        if "border" not in applied:
                            applied.append("border")
                    count += 1
            wb.save(tmp_path)
        finally:
            wb.close()
        return {"range": range, "cells_formatted": count, "applied": applied}

    try:
        summary = _atomic_write(target_path, kind, _do_write)
    except Exception as exc:
        return _write_error(kind, f"write failed: {exc.__class__.__name__}: {exc}", resolved)
    return _write_result(resolved, target_path, kind, summary)


def range_iter(lo: int, hi: int):
    """Inclusive integer range — `range(lo, hi+1)` shorthand."""
    return range(lo, hi + 1)


@tool("office.merge_cells")
def office_merge_cells(
    path: str,
    sheet: str | None,
    range: str,
) -> dict[str, Any]:
    """Merge a rectangular range into a single cell.

    Writes to `<basename>.edited.xlsx`; original preserved.

    Args:
        path: .xlsx to edit.
        sheet: sheet name (None = active).
        range: A1-style range like "A1:C1" (header row span) or
               "A1:B4" (vertical+horizontal merge).

    Returns:
        dict with `ok`, `original_path`, `edited_path`, `kind`,
        `summary={range}`.
    """
    resolved, kind, err = _validate_path(path)
    if err is not None:
        return _write_error(None, err)
    assert resolved is not None and kind is not None
    if kind != "excel":
        return _write_error(kind, "office.merge_cells requires a .xlsx file", resolved)
    if ":" not in range:
        return _write_error(
            kind, "merge_cells requires a range like 'A1:C1', not a single cell", resolved
        )

    target_path = _ensure_copy_for_write(resolved)

    def _do_write(tmp_path: str) -> dict[str, Any]:
        wb = load_workbook(target_path)
        try:
            ws = _resolve_sheet(wb, sheet)
            ws.merge_cells(range)
            wb.save(tmp_path)
        finally:
            wb.close()
        return {"range": range}

    try:
        summary = _atomic_write(target_path, kind, _do_write)
    except Exception as exc:
        return _write_error(kind, f"write failed: {exc.__class__.__name__}: {exc}", resolved)
    return _write_result(resolved, target_path, kind, summary)


@tool("office.add_row")
def office_add_row(
    path: str,
    sheet: str | None,
    values: list[Any],
    position: str | int = "append",
) -> dict[str, Any]:
    """Insert a row into a sheet.

    Writes to `<basename>.edited.xlsx`; original preserved.

    Args:
        path: .xlsx to edit.
        sheet: sheet name (None = active).
        values: list of cell values; the row is filled left-to-right
                from column A.
        position: "append" (default — adds after the last used row)
                  or an integer row number (1-indexed) to insert BEFORE
                  that row, shifting subsequent rows down.

    Returns:
        dict with `ok`, `original_path`, `edited_path`, `kind`,
        `summary={row, cells_written}`.
    """
    resolved, kind, err = _validate_path(path)
    if err is not None:
        return _write_error(None, err)
    assert resolved is not None and kind is not None
    if kind != "excel":
        return _write_error(kind, "office.add_row requires a .xlsx file", resolved)
    if not isinstance(values, list):
        return _write_error(kind, "values must be a list", resolved)
    if position != "append" and not isinstance(position, int):
        return _write_error(kind, "position must be 'append' or an integer row number", resolved)

    target_path = _ensure_copy_for_write(resolved)

    def _do_write(tmp_path: str) -> dict[str, Any]:
        wb = load_workbook(target_path)
        try:
            ws = _resolve_sheet(wb, sheet)
            if position == "append":
                ws.append(values)
                row_idx = ws.max_row
            else:
                row_idx = int(position)
                if row_idx < 1:
                    raise ValueError(f"row position must be >= 1, got {row_idx}")
                ws.insert_rows(row_idx)
                for i, v in enumerate(values, start=1):
                    ws.cell(row=row_idx, column=i, value=v)
            wb.save(tmp_path)
        finally:
            wb.close()
        return {"row": row_idx, "cells_written": len(values)}

    try:
        summary = _atomic_write(target_path, kind, _do_write)
    except Exception as exc:
        return _write_error(kind, f"write failed: {exc.__class__.__name__}: {exc}", resolved)
    return _write_result(resolved, target_path, kind, summary)


@tool("office.read_range")
def office_read_range(
    path: str,
    sheet: str | None,
    range: str,
) -> dict[str, Any]:
    """Read a specific cell range from a .xlsx as structured JSON.

    This is the precise complement to `office.read` (which dumps the
    whole workbook as markdown). Use this when you only need a subset
    — much smaller LLM context spend, and you get raw types (int /
    float / str / bool / None) plus the formula text for formula
    cells.

    Args:
        path: Runtime `/attachments/...` route or workspace .xlsx path.
        sheet: sheet name (None = active).
        range: A1-style range like "A1:C10" or single cell "B5".

    Returns:
        dict with:
          ok
          path, kind
          sheet (resolved sheet name)
          range (echoed back)
          values   — 2D list, computed values (formulas → cached result)
          formulas — 2D list, formula text where present, else None
    """
    source, kind, err = _resolve_read_source(path)
    if err is not None:
        return {"ok": "false", "error": err}
    assert source is not None and kind is not None
    reported_path = path if isinstance(source, bytes) else source
    if kind != "excel":
        return {
            "ok": "false",
            "error": "office.read_range requires a .xlsx file",
            "path": reported_path,
        }
    try:
        # Two passes: data_only=True for cached values, data_only=False
        # for formula text. openpyxl can't give both in one open.
        wb_values = load_workbook(
            io.BytesIO(source) if isinstance(source, bytes) else source,
            read_only=True,
            data_only=True,
        )
        wb_formulas = load_workbook(
            io.BytesIO(source) if isinstance(source, bytes) else source,
            read_only=True,
            data_only=False,
        )
    except Exception as exc:
        return {
            "ok": "false",
            "path": reported_path,
            "kind": kind,
            "error": f"failed to open .xlsx: {exc.__class__.__name__}: {exc}",
        }
    try:
        ws_values = _resolve_sheet(wb_values, sheet)
        ws_formulas = _resolve_sheet(wb_formulas, sheet)
        min_row, min_col, max_row, max_col = _parse_range(range)
        values_grid: list[list[Any]] = []
        formulas_grid: list[list[str | None]] = []
        for r in range_iter(min_row, max_row):
            v_row: list[Any] = []
            f_row: list[str | None] = []
            for c in range_iter(min_col, max_col):
                v_row.append(ws_values.cell(row=r, column=c).value)
                f_cell = ws_formulas.cell(row=r, column=c).value
                f_row.append(f_cell if isinstance(f_cell, str) and f_cell.startswith("=") else None)
            values_grid.append(v_row)
            formulas_grid.append(f_row)
    except Exception as exc:
        wb_values.close()
        wb_formulas.close()
        return {
            "ok": "false",
            "path": reported_path,
            "kind": kind,
            "error": f"failed to read range {range!r}: {exc.__class__.__name__}: {exc}",
        }
    sheet_name = ws_values.title
    wb_values.close()
    wb_formulas.close()
    return {
        "ok": "true",
        "path": reported_path,
        "kind": kind,
        "sheet": sheet_name,
        "range": range,
        "values": values_grid,
        "formulas": formulas_grid,
    }
